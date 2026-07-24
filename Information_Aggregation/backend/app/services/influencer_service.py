from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload

from app.models import Influencer, InfluencerProfile, InfluencerTag, User
from app.schemas import ImportResult, InfluencerCreate, InfluencerFilter, InfluencerUpdate
from app.utils.access_control import can_view_full_library, influencer_ids_for_user

PLATFORM_MAP = {
    "抖音": "douyin",
    "douyin": "douyin",
    "小红书": "xiaohongshu",
    "xiaohongshu": "xiaohongshu",
    "快手": "kuaishou",
    "kuaishou": "kuaishou",
    "微信": "wechat",
    "wechat": "wechat",
}

SOURCE_MAP = {
    "星图": "xingtu",
    "xingtu": "xingtu",
    "蒲公英": "pugongying",
    "pugongying": "pugongying",
    "互选": "huxuan",
    "huxuan": "huxuan",
    "手动": "manual",
    "manual": "manual",
}

IMPORT_COLUMNS = {
    "平台": "platform",
    "达人ID": "platform_uid",
    "达人昵称": "nickname",
    "昵称": "nickname",
    "粉丝量": "follower_count",
    "粉丝数": "follower_count",
    "来源": "source",
    "头像链接": "avatar_url",
    "主页链接": "profile_url",
    "互动率": "engagement_rate",
}


def _normalize_platform(value: Any) -> str | None:
    if value is None:
        return None
    return PLATFORM_MAP.get(str(value).strip().lower(), str(value).strip().lower())


def _normalize_source(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return SOURCE_MAP.get(text, SOURCE_MAP.get(text.lower(), text.lower()))


def _parse_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace(",", "").replace("万", "0000").strip()
    try:
        return int(float(text))
    except ValueError:
        return 0


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


class InfluencerService:
    @staticmethod
    def get_by_id(db: Session, influencer_id: int) -> Influencer | None:
        return (
            db.query(Influencer)
            .options(
                joinedload(Influencer.tags).joinedload(InfluencerTag.tag),
                joinedload(Influencer.profile),
                joinedload(Influencer.agency),
            )
            .filter(Influencer.id == influencer_id)
            .first()
        )

    @staticmethod
    def get_by_platform_uid(db: Session, platform: str, platform_uid: str) -> Influencer | None:
        return (
            db.query(Influencer)
            .filter(Influencer.platform == platform, Influencer.platform_uid == platform_uid)
            .first()
        )

    @staticmethod
    def list_influencers(
        db: Session,
        filters: InfluencerFilter,
        page: int,
        page_size: int,
        viewer: User | None = None,
    ) -> tuple[list[Influencer], int]:
        query = db.query(Influencer)

        if viewer is not None and not can_view_full_library(viewer):
            allowed_ids = influencer_ids_for_user(db, viewer.id)
            if not allowed_ids:
                return [], 0
            query = query.filter(Influencer.id.in_(allowed_ids))

        if filters.platform:
            query = query.filter(Influencer.platform == filters.platform)
        if filters.source:
            query = query.filter(Influencer.source == filters.source)
        if filters.status is not None:
            query = query.filter(Influencer.status == filters.status)
        if filters.follower_min is not None:
            query = query.filter(Influencer.follower_count >= filters.follower_min)
        if filters.follower_max is not None:
            query = query.filter(Influencer.follower_count <= filters.follower_max)
        if filters.keyword:
            keyword = f"%{filters.keyword}%"
            from sqlalchemy import or_, cast, String, exists

            from app.models import InfluencerProfile

            profile_hit = exists().where(
                InfluencerProfile.influencer_id == Influencer.id,
                or_(
                    cast(InfluencerProfile.cooperation_policy, String).like(keyword),
                    cast(InfluencerProfile.internal_notes, String).like(keyword),
                    cast(InfluencerProfile.shooting_style, String).like(keyword),
                    cast(InfluencerProfile.persona_traits, String).like(keyword),
                    cast(InfluencerProfile.contact_info, String).like(keyword),
                ),
            )
            query = query.filter(
                or_(
                    Influencer.nickname.like(keyword),
                    Influencer.platform_uid.like(keyword),
                    profile_hit,
                )
            )
        if filters.tag_ids:
            query = query.join(InfluencerTag).filter(InfluencerTag.tag_id.in_(filters.tag_ids))
        if filters.agency_id is not None:
            query = query.filter(Influencer.agency_id == filters.agency_id)

        total = query.distinct().count()
        items = (
            query.distinct()
            .options(
                joinedload(Influencer.tags).joinedload(InfluencerTag.tag),
                joinedload(Influencer.profile),
                joinedload(Influencer.agency),
            )
            .order_by(Influencer.updated_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return items, total

    @staticmethod
    def create(db: Session, data: InfluencerCreate) -> Influencer:
        existing = InfluencerService.get_by_platform_uid(db, data.platform, data.platform_uid)
        if existing:
            raise ValueError(f"达人已存在: {data.platform}/{data.platform_uid}")

        influencer = Influencer(**data.model_dump())
        db.add(influencer)
        db.commit()
        db.refresh(influencer)
        return influencer

    @staticmethod
    def update(db: Session, influencer: Influencer, data: InfluencerUpdate) -> Influencer:
        update_data = data.model_dump(exclude_unset=True, exclude={"profile"})
        for key, value in update_data.items():
            setattr(influencer, key, value)

        if data.profile is not None:
            profile = influencer.profile
            if profile is None:
                profile = InfluencerProfile(influencer_id=influencer.id)
                db.add(profile)
            profile_data = data.profile.model_dump(exclude_unset=True)
            for key, value in profile_data.items():
                setattr(profile, key, value)

        db.commit()
        db.refresh(influencer)
        return influencer

    @staticmethod
    def delete(db: Session, influencer: Influencer) -> None:
        db.delete(influencer)
        db.commit()

    @staticmethod
    def import_from_excel(db: Session, file_content: bytes) -> ImportResult:
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return ImportResult(total=0, success=0, failed=0, errors=["Excel 文件无数据"])

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map: dict[int, str] = {}
        for idx, header in enumerate(headers):
            if header in IMPORT_COLUMNS:
                col_map[idx] = IMPORT_COLUMNS[header]

        if "platform" not in col_map.values() or "platform_uid" not in col_map.values():
            return ImportResult(
                total=0,
                success=0,
                failed=0,
                errors=["缺少必填列：平台、达人ID"],
            )

        success = 0
        failed = 0
        errors: list[str] = []

        for row_num, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue

            record: dict[str, Any] = {}
            for idx, field in col_map.items():
                if idx < len(row):
                    record[field] = row[idx]

            platform = _normalize_platform(record.get("platform"))
            platform_uid = str(record.get("platform_uid", "")).strip()

            if not platform or not platform_uid:
                failed += 1
                errors.append(f"第{row_num}行: 平台或达人ID为空")
                continue

            payload = InfluencerCreate(
                platform=platform,
                platform_uid=platform_uid,
                nickname=str(record.get("nickname") or "").strip() or None,
                avatar_url=str(record.get("avatar_url") or "").strip() or None,
                profile_url=str(record.get("profile_url") or "").strip() or None,
                follower_count=_parse_int(record.get("follower_count")),
                engagement_rate=_parse_float(record.get("engagement_rate")),
                source=_normalize_source(record.get("source")) or "manual",
            )

            existing = InfluencerService.get_by_platform_uid(db, platform, platform_uid)
            try:
                if existing:
                    InfluencerService.update(
                        db,
                        existing,
                        InfluencerUpdate(**payload.model_dump(exclude={"platform", "platform_uid"})),
                    )
                else:
                    InfluencerService.create(db, payload)
                success += 1
            except Exception as exc:
                failed += 1
                errors.append(f"第{row_num}行: {exc}")

        total = success + failed
        return ImportResult(total=total, success=success, failed=failed, errors=errors[:20])
