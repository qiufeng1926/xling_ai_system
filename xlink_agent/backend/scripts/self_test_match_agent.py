"""商单筛库 / 隔离 自测（改代码后应跑通）。

用法（在 xlink_agent/backend 下）:
  set PYTHONPATH=.
  python scripts/self_test_match_agent.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import httpx

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

AGENT = "http://127.0.0.1:8003"
PORTAL = "http://127.0.0.1:8000"


def _client(**kwargs) -> httpx.Client:
    # Windows 系统代理可能导致本机 127.0.0.1 被错误代理成 502
    kwargs.setdefault("trust_env", False)
    kwargs.setdefault("timeout", 30.0)
    return httpx.Client(**kwargs)


def _fail(msg: str) -> None:
    print(f"[FAIL] {msg}")
    raise SystemExit(1)


def _ok(msg: str) -> None:
    print(f"[OK] {msg}")


def login_portal() -> str:
    candidates = [
        ("qiufengai", "qfai12@@"),
        ("admin", "admin123"),
        ("admin", "change-me-on-first-run"),
    ]
    with _client(timeout=15.0) as c:
        for user, pwd in candidates:
            r = c.post(
                f"{PORTAL}/api/v1/auth/login",
                data={"username": user, "password": pwd},
            )
            if r.status_code != 200:
                continue
            body = r.json()
            token = body.get("access_token")
            if not token and isinstance(body.get("data"), dict):
                token = body["data"].get("access_token")
            if token:
                _ok(f"portal login as {user}")
                return str(token)
    _fail(f"portal login failed（请确认 :8000 可用）。last={r.status_code} {r.text[:200]}")


def test_openapi_has_match() -> None:
    with _client(timeout=15.0) as c:
        r = c.get(f"{AGENT}/openapi.json")
        if r.status_code != 200:
            _fail(f"openapi {r.status_code}")
        paths = set((r.json().get("paths") or {}).keys())
    need = {
        "/api/agent/v1/match/conversations",
        "/api/agent/v1/match/conversations/{conversation_id}/chat",
        "/api/agent/v1/match/conversations/{conversation_id}/messages",
        "/api/agent/v1/match/conversations/{conversation_id}/export",
    }
    missing = need - paths
    if missing:
        _fail(f"openapi 缺少 match 路由: {missing}（需重启 :8003）")
    _ok("openapi 含商单筛库路由")


def test_match_crud(token: str) -> int:
    headers = {"Authorization": f"Bearer {token}"}
    with _client(timeout=30.0) as c:
        r = c.get(f"{AGENT}/api/agent/v1/match/conversations", headers=headers)
        if r.status_code != 200:
            _fail(f"GET match/conversations -> {r.status_code} {r.text[:300]}")
        _ok(f"list match conversations ({len(r.json().get('items') or [])})")

        r = c.post(
            f"{AGENT}/api/agent/v1/match/conversations",
            headers=headers,
            json={"title": "自测商单筛库"},
        )
        if r.status_code != 200:
            _fail(f"POST match/conversations -> {r.status_code} {r.text[:300]}")
        cid = int(r.json()["id"])
        if r.json().get("skill_slug") != "influencer-match":
            _fail(f"skill_slug 应为 influencer-match, got {r.json().get('skill_slug')}")
        _ok(f"create match conversation id={cid}")

        # 通用接口不得创建筛库会话
        r = c.post(
            f"{AGENT}/api/agent/v1/conversations",
            headers=headers,
            json={"title": "非法", "skill_slug": "influencer-match"},
        )
        if r.status_code == 200:
            _fail("通用 conversations 不应允许 skill_slug=influencer-match")
        _ok(f"通用接口拒绝筛库 slug ({r.status_code})")

        # 通用列表不应混入筛库会话（默认过滤）
        r = c.get(f"{AGENT}/api/agent/v1/conversations", headers=headers)
        if r.status_code != 200:
            _fail(f"list general conversations {r.status_code}")
        for it in r.json().get("items") or []:
            if it.get("id") == cid:
                _fail("通用会话列表泄漏了商单筛库会话")
        _ok("通用会话列表未泄漏筛库会话")
        return cid


def test_isolation_tools() -> None:
    from agent.orchestrator import KNOWN_TOOLS, _active_skills, _merge_tools
    from db.session import SessionLocal

    if any(str(t).startswith("influencer_") for t in KNOWN_TOOLS):
        _fail("通用 KNOWN_TOOLS 仍含 influencer_*")
    if "call_influencer_match" not in KNOWN_TOOLS:
        _fail("通用 KNOWN_TOOLS 缺少 call_influencer_match")
    db = SessionLocal()
    try:
        tools, _ = _merge_tools(_active_skills(db, 1))
    finally:
        db.close()
    if any(str(t).startswith("influencer_") for t in tools):
        _fail("通用 merge_tools 泄漏 influencer_*")
    if "call_influencer_match" not in tools:
        _fail("通用 merge_tools 缺少 call_influencer_match")
    if "web_search" not in tools:
        _fail("通用 merge_tools 缺少 web_search")
    _ok("工具隔离：通用可 call_influencer_match，无 influencer_*")


def test_runtime_blocks_direct_influencer() -> None:
    import asyncio

    from db.session import SessionLocal
    from tools.runtime import execute_tool

    db = SessionLocal()

    async def _run():
        return await execute_tool(
            "influencer_search",
            {"platform": "douyin"},
            db=db,
            user_id=1,
            conversation_id=1,
        )

    try:
        result = asyncio.run(_run())
    finally:
        db.close()
    if result.get("ok") is not False:
        _fail(f"通用 runtime 应拦截 influencer_search, got {result}")
    _ok("通用 runtime 拦截直接达人库工具")


def test_grounding() -> None:
    from match_agent.grounding import (
        build_grounded_answer,
        build_influencer_cards,
        ingest_observation_catalog,
    )

    cat: dict = {}
    ingest_observation_catalog(
        cat,
        {
            "ok": True,
            "items": [
                {
                    "id": 9,
                    "nickname": "真实达人",
                    "platform": "douyin",
                    "platform_uid": "uid9",
                    "follower_count": 1000,
                    "avatar_url": "https://example.com/a.png",
                    "contact": {"phone": "13800000009", "wechat": "wx9"},
                    "cooperation_policy": "政策A",
                    "shooting_style": ["口播"],
                    "persona_traits": ["亲和"],
                    "tags": ["探店"],
                }
            ],
        },
    )
    ans = build_grounded_answer(cat, brief="测试商单", draft="我编造假达人张三丰")
    if "张三丰" in ans:
        _fail("grounded 终稿混入了模型草稿编造")
    if "筛选出 1 位" not in ans and "1 位达人" not in ans:
        _fail(f"grounded 总起异常: {ans[:200]}")
    cards = build_influencer_cards(cat)
    if len(cards) != 1 or cards[0]["id"] != 9:
        _fail(f"cards 异常: {cards}")
    if cards[0].get("detail_path") != "/influencer/influencers/9":
        _fail(f"detail_path 异常: {cards[0]}")
    if cards[0].get("contact", {}).get("phone") != "13800000009":
        _fail("card 缺少联系方式")
    _ok("grounded 总起 + 卡片载荷仅含库记录")


def test_search_sanitization(token: str) -> None:
    """坏 platform / 假 tag_ids 不得把检索打成永久 0；keyword 应能命中标签名。"""
    import asyncio

    from tools.portal_context import reset_portal_bearer, set_portal_bearer
    from tools.influencer_tools import (
        _normalize_platform,
        _clean_tag_ids,
        influencer_search,
    )

    plat, warn = _normalize_platform("douyin|xiaohongshu")
    if plat is not None:
        _fail(f"多值 platform 应被忽略, got {plat}")
    if not warn:
        _fail("多值 platform 应有 warning")
    plat2, _ = _normalize_platform("抖音")
    if plat2 != "douyin":
        _fail(f"抖音应归一为 douyin, got {plat2}")
    ids, tw = _clean_tag_ids([123, 456, 9])
    if 123 in ids or 456 in ids:
        _fail(f"示例 tag_ids 应丢弃: {ids}")
    if 9 not in ids:
        _fail(f"真实 tag_id=9 应保留: {ids}")
    _ok("platform/tag_ids 规范化")

    tok = set_portal_bearer(token)

    async def _run():
        # 复现线上坏参：应自动放宽并仍能检出粉丝达标达人
        bad = await influencer_search(
            {
                "platform": "douyin|xiaohongshu",
                "follower_min": 50000,
                "tag_ids": [123, 456],
                "page_size": 10,
            }
        )
        # keyword=游戏 应能通过标签名命中（门户侧已含 tag LIKE）
        by_kw = await influencer_search(
            {
                "platform": "douyin",
                "follower_min": 50000,
                "keyword": "游戏",
                "page_size": 10,
            }
        )
        return bad, by_kw

    try:
        bad, by_kw = asyncio.run(_run())
    finally:
        reset_portal_bearer(tok)

    if not bad.get("ok"):
        _fail(f"坏参 search 失败: {bad}")
    if int(bad.get("count") or 0) <= 0 and int(bad.get("total") or 0) <= 0:
        _fail(f"坏参自动放宽后仍 0 人: {bad}")
    _ok(f"坏参自动放宽命中 count={bad.get('count')} widen={ (bad.get('applied') or {}).get('widen') }")

    if not by_kw.get("ok"):
        _fail(f"keyword=游戏 search 失败: {by_kw}")
    if int(by_kw.get("count") or 0) <= 0 and int(by_kw.get("total") or 0) <= 0:
        _fail(f"keyword=游戏 应命中标签达人，got 0: {by_kw}")
    _ok(f"keyword=游戏 命中 count={by_kw.get('count')} total={by_kw.get('total')}")


def test_match_messages(token: str, cid: int) -> None:
    headers = {"Authorization": f"Bearer {token}"}
    with _client(timeout=15.0) as c:
        r = c.get(f"{AGENT}/api/agent/v1/match/conversations/{cid}/messages", headers=headers)
        if r.status_code != 200:
            _fail(f"messages {r.status_code} {r.text[:200]}")
        _ok("match messages 可读")


def test_history_cards_and_export(token: str, cid: int) -> None:
    """历史回看卡片 + 一键导出闭环（不依赖 LLM）。"""
    import json

    from openpyxl import load_workbook

    from db.models import Conversation, Message
    from db.session import SessionLocal
    from match_agent.grounding import build_influencer_cards, ingest_observation_catalog

    headers = {"Authorization": f"Bearer {token}"}
    db = SessionLocal()
    try:
        conv = db.get(Conversation, cid)
        if not conv:
            _fail(f"conversation {cid} 不存在")
        cat: dict = {}
        ingest_observation_catalog(
            cat,
            {
                "ok": True,
                "items": [
                    {
                        "id": 101,
                        "nickname": "导出测试达人",
                        "platform": "douyin",
                        "platform_uid": "uid101",
                        "follower_count": 52000,
                        "agency_name": "测试MCN",
                        "tags": ["探店"],
                        "persona_traits": ["亲和"],
                        "shooting_style": ["口播"],
                        "cooperation_policy": "可议",
                        "contact": {"phone": "13900000101", "wechat": "wx101"},
                        "match_score": 88,
                        "match_reasons": ["粉丝量达标"],
                    }
                ],
            },
        )
        cards = build_influencer_cards(cat)
        msg = Message(
            conversation_id=cid,
            user_id=conv.user_id,
            role="assistant",
            content="筛选出 1 位达人（自测）",
            metadata_json=json.dumps(
                {"mode": "influencer-match", "influencers": cards},
                ensure_ascii=False,
            ),
        )
        db.add(msg)
        db.commit()
        db.refresh(msg)
        mid = int(msg.id)
    finally:
        db.close()

    with _client(timeout=20.0) as c:
        r = c.get(f"{AGENT}/api/agent/v1/match/conversations/{cid}/messages", headers=headers)
        if r.status_code != 200:
            _fail(f"messages hydrate {r.status_code}")
        found = None
        for it in r.json().get("items") or []:
            if it.get("id") == mid:
                found = it
                break
        if not found:
            _fail("messages 未返回刚写入的助手消息")
        infl = found.get("influencers") or []
        if len(infl) != 1 or infl[0].get("id") != 101:
            _fail(f"历史卡片未 hydrate: {infl}")
        _ok("历史 messages 回看含 influencers 卡片")

        r = c.get(
            f"{AGENT}/api/agent/v1/match/conversations/{cid}/export",
            headers=headers,
        )
        if r.status_code != 200:
            _fail(f"export latest -> {r.status_code} {r.text[:200]}")
        ctype = r.headers.get("content-type", "")
        if "spreadsheetml" not in ctype and "octet-stream" not in ctype:
            _fail(f"export content-type 异常: {ctype}")
        if len(r.content) < 800:
            _fail(f"export 体积过小: {len(r.content)}")
        from io import BytesIO

        wb = load_workbook(BytesIO(r.content))
        rows = list(wb.active.iter_rows(values_only=True))
        if len(rows) < 2 or rows[1][2] != "导出测试达人":
            _fail(f"export xlsx 内容异常: {rows[:3]}")
        _ok("export 默认导出最近一轮 xlsx")

        r = c.get(
            f"{AGENT}/api/agent/v1/match/conversations/{cid}/export",
            headers=headers,
            params={"message_id": mid},
        )
        if r.status_code != 200:
            _fail(f"export by message_id -> {r.status_code}")
        _ok(f"export 指定 message_id={mid}")

        r = c.get(
            f"{AGENT}/api/agent/v1/match/conversations/{cid}/export",
            headers=headers,
            params={"message_id": 999999999},
        )
        if r.status_code != 404:
            _fail(f"无卡片 export 应 404, got {r.status_code}")
        _ok("无结果 export 返回 404")


def main() -> None:
    print("=== self_test_match_agent ===")
    with _client(timeout=10.0) as c:
        try:
            h = c.get(f"{AGENT}/health")
        except Exception as exc:
            _fail(f"agent :8003 不可达: {exc}")
        if h.status_code != 200:
            _fail(f"agent health {h.status_code}")
    _ok("agent health")

    test_openapi_has_match()
    test_isolation_tools()
    test_runtime_blocks_direct_influencer()
    test_grounding()
    token = login_portal()
    test_search_sanitization(token)
    cid = test_match_crud(token)
    test_match_messages(token, cid)
    test_history_cards_and_export(token, cid)
    print("=== ALL PASSED ===")


if __name__ == "__main__":
    main()
