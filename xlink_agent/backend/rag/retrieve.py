from __future__ import annotations

import re
import uuid
from pathlib import Path

from sqlalchemy.orm import Session

from agent.model_router import get_chat_model
from db.models import KnowledgeDocument, RetrievalLog
from rag.qdrant_client import search_vectors, upsert_chunks
from utils.logger import get_logger

logger = get_logger("rag")


def chunk_text(text: str, size: int = 800, overlap: int = 100) -> list[str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    chunks: list[str] = []
    i = 0
    while i < len(text):
        chunks.append(text[i : i + size])
        i += max(size - overlap, 1)
    return chunks


def extract_text_from_file(path: Path, mime: str, filename: str) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".txt", ".html", ".htm"} or mime.startswith("text/"):
        raw = path.read_text(encoding="utf-8", errors="ignore")
        if suffix in {".html", ".htm"}:
            from bs4 import BeautifulSoup

            return BeautifulSoup(raw, "html.parser").get_text("\n")
        return raw
    if suffix == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        return "\n".join((p.extract_text() or "") for p in reader.pages)
    if suffix in {".docx"}:
        from docx import Document

        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)
    if suffix in {".xlsx", ".xls"}:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)
    return path.read_text(encoding="utf-8", errors="ignore")


async def ingest_document(db: Session, doc: KnowledgeDocument, kb_kind: str, owner_user_id: int | None) -> None:
    path = Path(doc.storage_path)
    try:
        text = extract_text_from_file(path, doc.mime, doc.filename)
        chunks = chunk_text(text)
        if not chunks:
            doc.status = "failed"
            doc.error_message = "未能提取有效文本"
            db.commit()
            return
        model = get_chat_model()
        vectors = await model.embed(chunks)
        dim = len(vectors[0]) if vectors else 64
        points = []
        for i, (chunk, vec) in enumerate(zip(chunks, vectors)):
            pid = str(uuid.uuid5(uuid.NAMESPACE_URL, f"{doc.id}:{i}"))
            # Qdrant 需要 unsigned int 或 UUID；使用 UUID 字符串
            points.append(
                {
                    "id": pid,
                    "vector": vec,
                    "payload": {
                        "text": chunk,
                        "doc_id": doc.id,
                        "kb_id": doc.kb_id,
                        "user_id": owner_user_id if kb_kind == "private" else 0,
                        "kind": kb_kind,
                        "filename": doc.filename,
                    },
                }
            )
        ok = upsert_chunks(points, dim)
        doc.status = "ready" if ok else "ready_mysql_only"
        if not ok:
            # 降级：把全文塞进 error_message 字段不合适；标记 ready_mysql_only 用文件再检索
            doc.error_message = "Qdrant 不可用，已保存原文供降级检索"
        db.commit()
    except Exception as exc:
        logger.exception("入库失败 doc=%s", doc.id)
        doc.status = "failed"
        doc.error_message = str(exc)[:1000]
        db.commit()


async def search_knowledge(db: Session, user_id: int, query: str, top_k: int = 5) -> list[dict]:
    model = get_chat_model()
    vec = (await model.embed([query]))[0]
    hits = search_vectors(vec, user_id=user_id, top_k=top_k)
    if not hits:
        # MySQL 降级：按文件名/状态 ready 的文档做简单包含匹配
        docs = (
            db.query(KnowledgeDocument)
            .filter(KnowledgeDocument.status.in_(["ready", "ready_mysql_only"]))
            .all()
        )
        for d in docs:
            # 权限：全局或本人
            from db.models import KnowledgeBase

            kb = db.get(KnowledgeBase, d.kb_id)
            if not kb:
                continue
            if kb.kind == "private" and kb.owner_user_id != user_id:
                continue
            try:
                text = extract_text_from_file(Path(d.storage_path), d.mime, d.filename)
            except Exception:
                continue
            if query.lower() in text.lower() or query in d.filename:
                hits.append(
                    {
                        "score": 0.5,
                        "text": text[:500],
                        "doc_id": d.id,
                        "kb_id": d.kb_id,
                        "filename": d.filename,
                    }
                )
            if len(hits) >= top_k:
                break

    db.add(
        RetrievalLog(
            user_id=user_id,
            conversation_id=None,
            query=query,
            hits_json=__import__("json").dumps(hits, ensure_ascii=False),
        )
    )
    db.commit()
    return hits
